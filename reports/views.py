import json
from bson import ObjectId
from datetime import datetime
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from report_webapp.utils import (reports, plans, kss, remarks,
                                 leaks, protocols, orders, authenticate_user,
                                 users, faults, reliability)
from django.http import JsonResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from reports.utils import parse_date_to_dmy, parse_departments


# Добавим mapping для преобразования названий служб
DEPARTMENT_MAPPING = {
    'начальник гкс': ['ГКС'],
    'начальники кс': ['КС-1,4', 'КС-2,3', 'КС-5,6', 'КС-7,8', 'КС-9,10'],
    'начальник кс-1,4': ['КС-1,4'],
    'начальник кс-2,3': ['КС-2,3'],
    'начальник кс-5,6': ['КС-5,6'],
    'начальник кс-7,8': ['КС-7,8'],
    'начальник кс-9,10': ['КС-9,10'],
    'начальник службы аимо': ['АиМО'],
    'начальник службы эвс': ['ЭВС'],
    'начальник службы лэс': ['ЛЭС'],
    'начальник службы сзк': ['СЗК'],
    'начальник службы связь': ['Связь'],
    'начальник службы впо': ['ВПО'],
    'инженера по ремонту': ['ГКС'],
    'инженер эого (техдиагностика)': ['ГКС'],
    'инженер по ремонту': ['ГКС'],
}


# Словарь для преобразования технических имен в читаемые
FIELD_NAMES_MAPPING = {
    # Общие поля
    'tasks': 'Задания на день',
    'faults': 'Замечания по оборудованию',

    # АПК
    'apk_total': 'Всего замечаний АПК I уровня',
    'apk_done': 'Устранено замечаний АПК I уровня',
    'apk_undone': 'Не устранено замечаний АПК I уровня',
    'apk_reason_undone': 'Причина неустранения АПК I уровня',
    'apk2_total': 'Всего замечаний АПК II уровня',
    'apk2_done': 'Устранено замечаний АПК II уровня',
    'apk2_undone': 'Не устранено замечаний АПК II уровня',
    'apk2_reason_undone': 'Причина неустранения АПК II уровня',

    # Утечки
    'leak_total': 'Обнаружено утечек',
    'leak_done': 'Устранено утечек',

    # Замечания
    'apk4_done': 'Устранено замечаний АПК IV уровня',
    'apk4_undone': 'Не устранено замечаний АПК IV уровня',
    'apk4_reason_undone': 'Причина неустранения АПК IV уровня',
    'ozp_done': 'Устранено замечаний ОЗП',
    'ozp_undone': 'Не устранено замечаний ОЗП',
    'ozp_reason_undone': 'Причина неустранения ОЗП',
    'gaz_done': 'Устранено замечаний Газнадзора',
    'gaz_undone': 'Не устранено замечаний Газнадзора',
    'gaz_reason_undone': 'Причина неустранения Газнадзора',
    'ros_done': 'Устранено замечаний Ростехнадзора',
    'ros_undone': 'Не устранено замечаний Ростехнадзора',
    'ros_reason_undone': 'Причина неустранения Ростехнадзора',

    # Рационализаторские предложения
    'rp_done': 'Подано РП',
    'rp_inwork': 'В работе РП',

    # ПАТ и ТУ
    'pat_done': 'Проведено ПАТ',
    'tu_done': 'Проведено ТУ',

    # КСС
    'kss_done': 'Выполнено КСС'
}


@csrf_exempt
def handle_report(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            current_year = datetime.now().year
            service = data.get('service')

            # Подготовка данных для сохранения
            report_data = {
                'department': service,
                'type': data.get('type'),
                'datetime': datetime.now(),
                'data': {}
            }

            # Удаляем служебные поля
            data.pop('service', None)
            data.pop('type', None)

            # Группируем данные по категориям
            categories = {
                'apk': ['apk_total', 'apk_done', 'apk_undone', 'apk_reason_undone'],
                'apk2': ['apk2_total', 'apk2_done', 'apk2_undone', 'apk2_reason_undone'],
                'leak': ['leak_total', 'leak_done'],
                'apk4': ['apk4_done', 'apk4_undone', 'apk4_reason_undone'],
                'ozp': ['ozp_done', 'ozp_undone', 'ozp_reason_undone'],
                'gaz': ['gaz_done', 'gaz_undone', 'gaz_reason_undone'],
                'ros': ['ros_done', 'ros_undone', 'ros_reason_undone'],
                'rp': ['rp_done', 'rp_inwork'],
                'pat': ['pat_done'],
                'tu': ['tu_done'],
                'kss': ['kss_done']
            }

            # Основные поля
            report_data['data']['tasks'] = data.get('task', '')
            report_data['data']['faults'] = data.get('faults', '')

            # Обрабатываем категории
            for category, fields in categories.items():
                category_data = {}
                for field in fields:
                    if field in data:
                        value = data[field]
                        # Для полей с причинами оставляем текст как есть
                        if 'reason' in field:
                            category_data[field] = str(value) if value is not None else ''
                        # Для числовых полей преобразуем в int
                        elif any(x in field for x in ['total', 'done', 'undone', 'inwork']):
                            try:
                                category_data[field] = int(value) if value else 0
                            except (ValueError, TypeError):
                                category_data[field] = 0
                        # Все остальные поля оставляем как есть
                        else:
                            category_data[field] = value

                if category_data:  # Добавляем только если есть данные
                    report_data['data'][category] = category_data

            # Сохраняем в MongoDB
            reports.insert_one(report_data)

            # Обновляем дополнительные коллекции
            update_related_collections(report_data, current_year)

            protocol_updates = {}
            for key, value in data.items():
                if key.startswith('protocol_') and value == 'on':
                    protocol_id = key.replace('protocol_', '')
                    protocol_updates[protocol_id] = datetime.now()

            if protocol_updates:
                for protocol_id, done_date in protocol_updates.items():
                    protocols.update_one(
                        {'_id': ObjectId(protocol_id)},
                        {'$set': {f'done.{service}': done_date}}
                    )

            return JsonResponse({'status': 'success', 'message': 'Данные успешно сохранены'})
        except Exception as e:
            return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': str(e)}))

    elif request.method == 'GET':
        try:
            service = request.GET.get('service')
            report_type = request.GET.get('type')  # daily или weekly
            limit = int(request.GET.get('limit', 1))  # Количество отчетов для загрузки
            skip = int(request.GET.get('skip', 0))  # Пропустить первые N отчетов

            if not service:
                return JsonResponse({'status': 'error', 'message': 'Не указана служба'}, status=400)

            query = {'department': service}
            if report_type:
                query['type'] = report_type

            # Получаем отчеты с пагинацией
            reports_cursor = reports.find(
                query,
                {'_id': 0, 'data': 1, 'datetime': 1, 'type': 1}
            ).sort('datetime', -1).skip(skip).limit(limit)

            reports_list = list(reports_cursor)
            for report in reports_list:
                report['datetime'] = report['datetime'].isoformat()

            # Получаем общее количество отчетов для навигации
            total_count = reports.count_documents(query)

            return JsonResponse({
                'status': 'success',
                'reports': reports_list,
                'total_count': total_count,
                'skip': skip,
                'limit': limit
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


def update_related_collections(report_data, year):
    """Обновление связанных коллекций (замечания, утечки, КСС)"""
    department = report_data['department']
    now = datetime.now()

    # Обработка утечек
    if 'leak' in report_data['data']:
        leak_total = report_data['data']['leak'].get('leak_total', 0)
        leak_done = report_data['data']['leak'].get('leak_done', 0)
        leaks.update_one(
            {'year': year, 'department': department},
            {
                '$inc': {'total': leak_total, 'done': leak_done},
                '$setOnInsert': {'datetime': now}
            },
            upsert=True
        )

    # Обработка КСС
    if 'kss' in report_data['data']:
        kss_done = report_data['data']['kss'].get('kss_done', 0)
        if kss_done > 0:
            kss.update_one(
                {'year': year},
                {
                    '$inc': {'total': kss_done},
                    '$setOnInsert': {'datetime': now}
                },
                upsert=True
            )

    # Обработка замечаний (ОЗП, Газнадзор, Ростехнадзор)
    for remark_type in ['ozp', 'gaz', 'ros', 'apk4']:
        if remark_type in report_data['data']:
            remark_done = report_data['data'][remark_type].get(f'{remark_type}_done', 0)
            if remark_done > 0:
                # Проверяем, есть ли запись для текущего года
                existing_remark = remarks.find_one({
                    'year': year,
                    'value': remark_type,
                    'department': department
                })
                if existing_remark:
                    # Обновляем существующую запись
                    remarks.update_one(
                        {'year': year, 'value': remark_type, 'department': department},
                        {
                            '$inc': {'done': remark_done},
                            '$setOnInsert': {'datetime': now}
                        },
                        upsert=True
                    )


@csrf_exempt
def handle_planning(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            department = data.get('service')
            year = int(data.get('year'))
            now = datetime.now()

            # Обработка замечаний (ОЗП, Газнадзор, Ростехнадзор)
            for remark_type in ['ozp', 'gaz', 'ros', 'apk4']:
                total = data.get(f'{remark_type}_total', 0)
                if total:
                    remark_data = {
                        'department': department,
                        'year': year,
                        'datetime': now,
                        'value': remark_type,
                        'total': int(total),
                        'done': 0  # Начальное значение выполненных
                    }
                    remarks.replace_one(
                        {'department': department, 'year': year, 'value': remark_type},
                        remark_data,
                        upsert=True
                    )

            # Обработка планов (РП, ПАТ, ТУ)
            for plan_type in ['rp', 'pat', 'tu']:
                total = data.get(f'{plan_type}_total', 0)
                if total:
                    quarters = {
                        '1': int(data.get(f'{plan_type}_q1', 0)),
                        '2': int(data.get(f'{plan_type}_q2', 0)),
                        '3': int(data.get(f'{plan_type}_q3', 0)),
                        '4': int(data.get(f'{plan_type}_q4', 0))
                    }

                    plan_data = {
                        'department': department,
                        'datetime': now,
                        'year': year,
                        'value': plan_type,
                        'total': int(total),
                        'quarters': quarters
                    }
                    plans.replace_one(
                        {'department': department, 'year': year, 'value': plan_type},
                        plan_data,
                        upsert=True
                    )
            return JsonResponse({'status': 'success', 'message': 'Данные планирования успешно сохранены'})
        except Exception as e:
            return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': str(e)}))
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


def get_reports(request):
    """Получение отчетов для отображения"""
    service = request.GET.get('service')
    if not service:
        return JsonResponse({'status': 'error', 'message': 'Не указана служба'}, status=400)

    try:
        # Получаем последний ежедневный отчет
        daily_report = reports.find_one(
            {'department': service, 'type': 'daily'},
            {'_id': 0, 'data': 1, 'datetime': 1, 'type': 1},
            sort=[('datetime', -1)]
        )

        # Получаем последний еженедельный отчет
        weekly_report = reports.find_one(
            {'department': service, 'type': 'weekly'},
            {'_id': 0, 'data': 1, 'datetime': 1, 'type': 1},
            sort=[('datetime', -1)]
        )

        reports_list = []
        if daily_report:
            daily_report['datetime'] = daily_report['datetime'].isoformat()
            reports_list.append(daily_report)
        if weekly_report:
            weekly_report['datetime'] = weekly_report['datetime'].isoformat()
            reports_list.append(weekly_report)

        return JsonResponse({'status': 'success', 'reports': reports_list})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def index(request):
    return render(request, 'report_webapp/index.html')


def view_data(request):
    """Страница просмотра данных"""
    services = [
        'КС-1,4', 'КС-2,3', 'КС-5,6', 'КС-7,8', 'КС-9,10',
        'ГКС', 'АиМО', 'ЭВС', 'ЛЭС', 'СЗК', 'Связь', 'ВПО'
    ]
    return render(request, 'report_webapp/index.html', {'services': services})


@csrf_exempt
def get_leaks(request):
    department = request.GET.get('department')
    year = request.GET.get('year')

    try:
        leaks_data = leaks.find_one(
            {'department': department, 'year': int(year)},
            {'_id': 0, 'total': 1, 'done': 1}
        )
        return JsonResponse({'status': 'success', 'total': leaks_data.get('total', 0), 'done': leaks_data.get('done', 0)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def get_kss(request):
    year = request.GET.get('year')

    try:
        kss_data = kss.find_one(
            {'year': int(year)},
            {'_id': 0, 'total': 1}
        )
        return JsonResponse({'status': 'success', 'total': kss_data.get('total', 0)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def get_remarks(request):
    department = request.GET.get('department')
    year = request.GET.get('year')

    try:
        remarks_list = list(remarks.find(
            {'department': department, 'year': int(year)},
            {'_id': 0, 'value': 1, 'total': 1, 'done': 1}
        ))
        return JsonResponse({
            'status': 'success',
            'remarks': remarks_list  # Теперь возвращаем массив в поле remarks
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def get_plans(request):
    try:
        department = request.GET.get('department')
        year = request.GET.get('year')

        plans_list = list(plans.find(
            {'department': department, 'year': int(year)},
            {'_id': 0, 'value': 1, 'total': 1, 'quarters': 1}
        ))

        return JsonResponse({
            'status': 'success',
            'plans': plans_list,
            'message': 'Plans data loaded successfully'
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'plans': []
        }, status=500)


@csrf_exempt
def handle_protocols(request):
    if request.method == 'GET':
        try:
            # Получаем только неархивированные протоколы
            queryset = list(protocols.find(
                {'archived': {'$ne': True}},
                {'_id': 1, 'date': 1, 'text': 1, 'archived': 1, 'done': 1,
                 'departments': 1, 'issue_date': 1, 'protocol_num': 1, 'protocol_name': 1, 'deadline': 1}
            ).sort('issue_date', -1))  # Сортируем по дате выхода (новые сверху)

            # Преобразуем ObjectId в строку и форматируем даты
            formatted_protocols = []
            for protocol in queryset:
                formatted = {
                    '_id': str(protocol['_id']),
                    'protocol_num': protocol.get('protocol_num', ''),
                    'protocol_name': protocol.get('protocol_name', ''),
                    'issue_date': protocol['issue_date'].isoformat() if 'issue_date' in protocol else '',
                    'deadline': protocol.get('deadline', ''),  # Текстовое поле
                    'departments': protocol['departments'],
                    'text': protocol['text'],
                    'done': {}
                }

                # Форматируем информацию о выполнении
                if 'done' in protocol:
                    for dept, date in protocol['done'].items():
                        if isinstance(date, datetime):
                            formatted['done'][dept] = date.isoformat()
                        else:
                            # Если дата уже в строковом формате (на всякий случай)
                            formatted['done'][dept] = date

                formatted_protocols.append(formatted)

            return JsonResponse({
                'status': 'success',
                'protocols': formatted_protocols
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            protocol_data = {
                'issue_date': datetime.fromisoformat(data['issue_date']),  # Новая дата выхода
                'protocol_num': data['protocol_num'],  # Номер протокола
                'protocol_name': data['protocol_name'],  # Название протокола
                'deadline': data['deadline'],  # Текстовый срок исполнения
                'text': data['text'],
                'departments': data['departments'],
                'archived': False,
                'created_at': datetime.now()
            }

            result = protocols.insert_one(protocol_data)
            return JsonResponse({
                'status': 'success',
                'message': 'Протокол успешно добавлен',
                'id': str(result.inserted_id)
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            protocol_id = request.path.split('/')[-2]  # Из URL /api/protocols/{id}/
            update_data = {
                'protocol_num': data.get('protocol_num'),
                'protocol_name': data.get('protocol_name'),
                'issue_date': datetime.fromisoformat(data.get('issue_date')),
                'deadline': data.get('deadline'),
                'text': data.get('text'),
                'departments': data.get('departments')
            }
            result = protocols.update_one(
                {'_id': ObjectId(protocol_id)},
                {'$set': update_data}
            )
            if result.modified_count:
                return JsonResponse({'status': 'success'})
            return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def archive_protocol(request, protocol_id):
    print(protocols.find_one({'_id': ObjectId(protocol_id)}))
    if request.method == 'POST':
        try:
            result = protocols.update_one(
                {'_id': ObjectId(protocol_id)},
                {'$set': {'archived': True, 'archived_at': datetime.now()}}
            )

            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Протокол архивирован'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Протокол не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def mark_protocol_done(request, protocol_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service = data.get('service')
            done_date = datetime.fromisoformat(data.get('done_date').replace('Z', '+00:00'))

            result = protocols.update_one(
                {'_id': ObjectId(protocol_id)},
                {'$set': {f'done.{service}': done_date}}
            )

            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Протокол обновлен'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Протокол не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def handle_orders(request):
    if request.method == 'GET':
        try:
            # Получаем только неархивированные распоряжения
            queryset = list(orders.find(
                {'archived': {'$ne': True}},
                {'_id': 1, 'date': 1, 'text': 1, 'archived': 1, 'done': 1, 'num': 1, 'departments': 1, 'issue_date': 1, 'deadline': 1}
            ).sort('issue_date', -1))  # Сортируем по дате выхода (новые сверху)

            # Преобразуем ObjectId в строку и форматируем даты
            formatted_orders = []
            for order in queryset:
                formatted = {
                    '_id': str(order['_id']),
                    'num': order['num'],
                    'issue_date': order['issue_date'].isoformat() if 'issue_date' in order else '',
                    'deadline': order.get('deadline', ''),  # Текстовое поле
                    'departments': order['departments'],
                    'text': order['text'],
                    'done': {}
                }
                # Форматируем информацию о выполнении
                if 'done' in order:
                    for dept, date in order['done'].items():
                        if isinstance(date, datetime):
                            formatted['done'][dept] = date.isoformat()
                        else:
                            formatted['done'][dept] = date
                formatted_orders.append(formatted)
            return JsonResponse({
                'status': 'success',
                'orders': formatted_orders
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_data = {
                'issue_date': datetime.fromisoformat(data['issue_date']),  # Новая дата выхода
                'deadline': data['deadline'],  # Текстовый срок исполнения
                'text': data['text'],
                'num': data['num'],
                'departments': data['departments'],
                'archived': False,
                'created_at': datetime.now()
            }
            result = orders.insert_one(order_data)
            return JsonResponse({
                'status': 'success',
                'message': 'Распоряжение (приказ) успешно добавлено',
                'id': str(result.inserted_id)
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def archive_order(request, order_id):
    if request.method == 'POST':
        try:
            result = orders.update_one(
                {'_id': ObjectId(order_id)},
                {'$set': {'archived': True, 'archived_at': datetime.now()}}
            )
            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Распоряжение архивировано'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Распоряжение не найдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def mark_order_done(request, order_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service = data.get('service')
            done_date = datetime.fromisoformat(data.get('done_date').replace('Z', '+00:00'))
            result = orders.update_one(
                {'_id': ObjectId(order_id)},
                {'$set': {f'done.{service}': done_date}}
            )
            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Протокол обновлен'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Протокол не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def authenticate_view(request):
    if request.method == "POST":
        body = json.loads(request.body)
        department = body.get("department")
        password = body.get("password")
        user = authenticate_user(department, password)
        if user:
            return JsonResponse({"token": "valid"}, status=200)
        else:
            return JsonResponse({"error": "Invalid credentials"}, status=401)


def departments_list(request):
    deps = users.distinct("department")
    return JsonResponse(list(deps), safe=False)


@csrf_exempt
def handle_faults(request):
    if request.method == 'GET':
        try:
            # Получаем только неархивированные распоряжения
            queryset = list(faults.find(
                {'archived': {'$ne': True}},
                {'_id': 1, 'date': 1, 'text': 1, 'archived': 1, 'is_done': 1, 'num': 1, 'department': 1, 'type': 1, 'date_done': 1}
            ).sort('date', 1))
            # Преобразуем ObjectId в строку и форматируем даты
            formatted_faults = []
            for fault in queryset:
                formatted = {
                    '_id': str(fault['_id']),
                    'date': fault['date'].isoformat(),
                    'department': fault['department'],
                    'type': fault['type'],
                    'text': fault['text'],
                    'is_done': fault['is_done'],
                    'date_done': fault['date_done']
                }
                formatted_faults.append(formatted)
            return JsonResponse({
                'status': 'success',
                'faults': formatted_faults
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            fault_data = {
                'date': datetime.fromisoformat(data['date']),
                'text': data['text'],
                'type': data['type'],
                'department': data['department'],
                'archived': False,
                'created_at': datetime.now(),
                'is_done': False,
                'date_done': datetime.now(),
            }
            result = faults.insert_one(fault_data)
            return JsonResponse({
                'status': 'success',
                'message': 'Замечание успешно добавлено',
                'id': str(result.inserted_id)
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def archive_fault(request, fault_id):
    if request.method == 'POST':
        try:
            result = faults.update_one(
                {'_id': ObjectId(fault_id)},
                {'$set': {'archived': True, 'archived_at': datetime.now()}}
            )
            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Распоряжение архивировано'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Распоряжение не найдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def mark_fault_done(request, fault_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service = data.get('service')
            done_date = datetime.fromisoformat(data.get('done_date').replace('Z', '+00:00'))
            result = faults.update_one(
                {'_id': ObjectId(fault_id)},
                {'$set': {f'is_done': True, 'date_done': done_date}}
            )
            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Замечание обновлено'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Замечание не найдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return HttpResponseBadRequest(json.dumps({'status': 'error', 'message': 'Неверный метод запроса'}))


@csrf_exempt
def handle_reliability(request):
    if request.method == 'GET':
        try:
            # Получаем только неархивированные мероприятия
            queryset = list(reliability.find(
                {'archived': {'$ne': True}},
                {'_id': 1, 'name': 1, 'date': 1, 'departments': 1,
                 'note': 1, 'archived': 1, 'done': 1}
            ).sort('deadline', 1))

            formatted_items = []
            for item in queryset:
                formatted = {
                    '_id': str(item['_id']),
                    'name': item['name'],
                    'date': item['date'],
                    'departments': item['departments'],
                    'note': item.get('note', ''),
                    'done': {}
                }

                if 'done' in item:
                    for dept, date in item['done'].items():
                        if isinstance(date, datetime):
                            formatted['done'][dept] = date.isoformat()
                        else:
                            formatted['done'][dept] = date

                formatted_items.append(formatted)

            return JsonResponse({
                'status': 'success',
                'items': formatted_items
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            reliability_data = {
                'name': data['name'],
                'date': data['date'],
                'departments': data['departments'],
                'note': data.get('note', ''),
                'archived': False,
                'created_at': datetime.now(),
                'done': {}
            }

            result = reliability.insert_one(reliability_data)
            return JsonResponse({
                'status': 'success',
                'message': 'Мероприятие успешно добавлено',
                'id': str(result.inserted_id)
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def archive_reliability(request, item_id):
    if request.method == 'POST':
        try:
            result = reliability.update_one(
                {'_id': ObjectId(item_id)},
                {'$set': {'archived': True, 'archived_at': datetime.now()}}
            )

            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Мероприятие архивировано'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Мероприятие не найдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def mark_reliability_done(request, item_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service = data.get('service')
            done_date = datetime.fromisoformat(data.get('done_date').replace('Z', '+00:00'))

            result = reliability.update_one(
                {'_id': ObjectId(item_id)},
                {'$set': {f'done.{service}': done_date}}
            )

            if result.modified_count == 1:
                return JsonResponse({'status': 'success', 'message': 'Мероприятие обновлено'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Мероприятие не найдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def upload_reliability_excel(request):
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                return JsonResponse({'status': 'error', 'message': 'Файл не найден'}, status=400)

            excel_file = request.FILES['excel_file']

            # Загружаем Excel файл
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active

            # Функция для поиска строки с заголовками
            def find_header_row(sheet):
                header_patterns = [
                    'наименование мероприятия',
                    'сроки реализации',
                    'ответственные',
                    'примечание',
                    'оборудование'
                ]

                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and any(pattern in str(cell_value).lower() for pattern in header_patterns):
                            return row
                return None

            # Находим строку с заголовками
            header_row = find_header_row(sheet)
            if not header_row:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Не удалось найти заголовки таблицы в файле'
                }, status=400)

            # Определяем столбцы по заголовкам
            headers = {}
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value:
                    header = str(cell_value).lower().strip()
                    if 'наименование мероприятия' in header:
                        headers['name'] = col
                    elif 'наименование/тип оборудования' in header or 'оборудование' in header:
                        headers['equipment'] = col
                    elif 'сроки реализации' in header or 'периодичность выполнения' in header:
                        headers['date'] = col
                    elif 'ответственные' in header:
                        headers['departments'] = col
                    elif 'примечание' in header:
                        headers['note'] = col

            if not all(key in headers for key in ['name', 'date', 'departments']):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Не найдены необходимые столбцы: Наименование мероприятия, Сроки реализации, Ответственные'
                }, status=400)

            imported_count = 0
            skipped_count = 0

            # Обрабатываем строки начиная со следующей после заголовков
            for row in range(header_row + 1, sheet.max_row + 1):
                name = sheet.cell(row=row, column=headers['name']).value
                equipment = sheet.cell(row=row, column=headers.get('equipment')).value if headers.get('equipment') else None
                date = sheet.cell(row=row, column=headers['date']).value
                departments_text = sheet.cell(row=row, column=headers['departments']).value
                note = sheet.cell(row=row, column=headers.get('note')).value if headers.get('note') else None

                # Пропускаем пустые строки (где нет названия мероприятия)
                if not name:
                    continue

                # Формируем полное название мероприятия
                full_name = str(name)
                if equipment:
                    full_name = f"{equipment}. {full_name}"

                # Преобразуем дату в строку в формате ДД.ММ.ГГГГ
                date_str = parse_date_to_dmy(date)

                # Парсим ответственные службы
                departments = parse_departments(str(departments_text) if departments_text else '')

                # Пропускаем если нет ответственных служб
                if not departments:
                    skipped_count += 1
                    continue

                # Создаем запись мероприятия
                reliability_data = {
                    'name': full_name,
                    'date': date_str,
                    'departments': departments,
                    'note': str(note) if note else '',
                    'archived': False,
                    'created_at': datetime.now(),
                    'done': {},
                    'source': 'excel_import'
                }

                # Проверяем, нет ли уже такого мероприятия
                existing = reliability.find_one({
                    'name': reliability_data['name'],
                    'date': reliability_data['date']
                })

                if not existing:
                    reliability.insert_one(reliability_data)
                    imported_count += 1
                else:
                    skipped_count += 1

            return JsonResponse({
                'status': 'success',
                'message': f'Успешно импортировано {imported_count} мероприятий, пропущено {skipped_count} (дубликаты или без служб)'
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Ошибка обработки файла: {str(e)}'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Неверный метод запроса'}, status=400)
